"""
GrappleMap → Excel exporter
Baixa GrappleMap.txt do GitHub e gera um .xlsx com todas as entradas
organizadas em abas por categoria, incluindo coordenadas 3D de cada articulação
dos 2 adversários (decodificadas do formato base62 do GrappleMap).
"""

import requests
import math
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuração ──────────────────────────────────────────────────────────────

URL = "https://raw.githubusercontent.com/Eelis/GrappleMap/master/GrappleMap.txt"
OUTPUT = "GrappleMap_v2.xlsx"

# Tags que indicam categoria de posição/sistema
POSITION_TAGS = {
    "standing", "mount", "side_control", "judo_side", "north_south",
    "half_guard", "quarter_guard", "three_quarter_guard", "full_guard",
    "closed_guard", "butterfly", "reverse_butterfly", "engaged_butterfly",
    "x_guard", "rubber_guard", "spider_guard", "inverted_guard",
    "back", "turtle", "footsies", "dogfight", "crucifix", "twister_side",
    "knee_on_belly", "s_mount", "three_quarter_mount", "crab_ride",
    "truck", "ashi", "lockdown", "deep_half", "z_guard",
    "knee_shield", "collar_tie", "clinch",
}

# Tags que indicam tipo de técnica
TECHNIQUE_TAGS = {
    "sweep", "pass", "pass_through", "pass_over", "pass_under",
    "pass_around", "pass_split", "pass_double_under",
    "takedown", "double_leg_takedown", "single_leg_takedown",
    "throw", "sacrifice_throw", "te_waza", "koshi_waza", "ashi_waza", "sutemi_waza",
    "armbar", "kimura", "omoplata", "triangle", "rear_naked_choke",
    "arm_choke", "arm_triangle", "guillotine", "darce", "toehold",
    "heel_hook", "knee_bar", "neck_crank", "shoulder_lock", "monoplata",
    "hip_bump", "bridge", "stand_up", "escape", "guard_pull", "guard_jump",
    "roll", "inversion", "back_step",
}

# Postura do lutador de cima
TOP_POSTURE_TAGS = {
    "top_kneeling", "top_on_side", "top_seated", "top_sitting",
    "top_supine", "top_airborne", "top_post_hand", "top_double_unders",
    "top_underhook", "top_overhook", "top_arm_pin", "top_posture_broken",
}

# Postura do lutador de baixo
BOTTOM_POSTURE_TAGS = {
    "bottom_supine", "bottom_prone", "bottom_seated", "bottom_kneeling",
    "bottom_on_side", "bottom_inverted", "bottom_turned_in",
    "bottom_turned_away", "bottom_post_hand", "bottom_post_elbow",
    "bottom_underhook", "bottom_overhook", "bottom_open_elbow",
    "bottom_double_unders",
}

# ── Decodificador 3D ─────────────────────────────────────────────────────────
#
# Formato do GrappleMap (persistence.cpp):
#   - Alfabeto base62: a-z (0-25), A-Z (26-51), 0-9 (52-61)
#   - Cada coordenada = 2 chars  →  valor = (c0*62 + c1) / 1000.0
#   - Offset: X armazenado como X+2, Z armazenado como Z+2 (Y sem offset)
#   - Ordem: player0 joints 0-22, depois player1 joints 0-22
#   - Cada frame = 4 linhas de 69 chars (sem o indent de 4 espaços)
#     = 276 chars = 2 players × 23 joints × 3 coords × 2 chars

BASE62 = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
BASE62_MAP = {c: i for i, c in enumerate(BASE62)}

# 23 articulações na ordem do GrappleMap
JOINTS = [
    "LeftToe",      "RightToe",
    "LeftHeel",     "RightHeel",
    "LeftAnkle",    "RightAnkle",
    "LeftKnee",     "RightKnee",
    "LeftHip",      "RightHip",
    "LeftShoulder", "RightShoulder",
    "LeftElbow",    "RightElbow",
    "LeftWrist",    "RightWrist",
    "LeftHand",     "RightHand",
    "LeftFingers",  "RightFingers",
    "Core",         "Neck",         "Head",
]

JOINTS_PT = [
    "Dedo_Pé_Esq",  "Dedo_Pé_Dir",
    "Calcanhar_Esq", "Calcanhar_Dir",
    "Tornozelo_Esq", "Tornozelo_Dir",
    "Joelho_Esq",    "Joelho_Dir",
    "Quadril_Esq",   "Quadril_Dir",
    "Ombro_Esq",     "Ombro_Dir",
    "Cotovelo_Esq",  "Cotovelo_Dir",
    "Pulso_Esq",     "Pulso_Dir",
    "Mão_Esq",       "Mão_Dir",
    "Dedos_Esq",     "Dedos_Dir",
    "Core",          "Pescoço",      "Cabeça",
]


def decode_frame(b64_lines: list[str]) -> dict | None:
    """
    Decodifica um frame (4 linhas de base62) em coordenadas 3D.
    Retorna {'player0': [(x,y,z), ...], 'player1': [(x,y,z), ...]}
    ou None se as linhas tiverem tamanho incorreto.
    """
    s = "".join(line.strip() for line in b64_lines)
    expected = 2 * len(JOINTS) * 3 * 2  # = 276
    if len(s) < expected:
        return None

    pos = 0

    def next_coord(offset: float) -> float:
        nonlocal pos
        c0 = BASE62_MAP.get(s[pos], 0); pos += 1
        c1 = BASE62_MAP.get(s[pos], 0); pos += 1
        return round((c0 * 62 + c1) / 1000.0 - offset, 4)

    result = {"player0": [], "player1": []}
    for player_key in ("player0", "player1"):
        for _ in JOINTS:
            x = next_coord(2.0)  # stored as X+2
            y = next_coord(0.0)  # no offset
            z = next_coord(2.0)  # stored as Z+2
            result[player_key].append((x, y, z))

    return result


def decode_all_frames(raw_b64_lines: list[str]) -> list[dict] | None:
    """Decodifica todos os frames de uma entrada (cada 4 linhas = 1 frame)."""
    if not raw_b64_lines or len(raw_b64_lines) % 4 != 0:
        return None
    frames = []
    for i in range(0, len(raw_b64_lines), 4):
        frame = decode_frame(raw_b64_lines[i:i+4])
        if frame:
            frames.append(frame)
    return frames if frames else None


def centroid(coords: list[tuple]) -> tuple:
    n = len(coords)
    return (
        round(sum(c[0] for c in coords) / n, 4),
        round(sum(c[1] for c in coords) / n, 4),
        round(sum(c[2] for c in coords) / n, 4),
    )


def player_height(coords: list[tuple]) -> float:
    """Altura do jogador = máximo Y menos mínimo Y."""
    ys = [c[1] for c in coords]
    return round(max(ys) - min(ys), 4)


def players_distance(p0_coords: list[tuple], p1_coords: list[tuple]) -> float:
    """Distância entre os centróides dos dois jogadores."""
    c0 = centroid(p0_coords)
    c1 = centroid(p1_coords)
    return round(math.sqrt(sum((a-b)**2 for a, b in zip(c0, c1))), 4)


# ── Parser ────────────────────────────────────────────────────────────────────

def parse_grapplemap(text: str) -> list[dict]:
    """
    Parseia o arquivo GrappleMap.txt e retorna uma lista de entradas.
    Cada entrada tem: name, name_clean, tags, b64_lines, frames, status,
                      pos_tags, tech_tags, top_posture, bottom_posture,
                      main_category, category_detail
    """
    lines = text.split("\n")
    entries = []
    i = 0

    while i < len(lines):
        line = lines[i]

        # Linha de nome: não começa com espaço nem com 'tags:'
        if line and not line.startswith("    ") and not line.startswith("tags:"):
            name_raw = line.strip()
            j = i + 1

            # Linha de tags
            if j < len(lines) and lines[j].startswith("tags:"):
                tags_raw = lines[j].replace("tags:", "").strip()
                tags = tags_raw.split() if tags_raw else []
                j += 1

                # Linhas de base62
                raw_lines = []
                while j < len(lines) and lines[j].startswith("    "):
                    raw_lines.append(lines[j])
                    j += 1

                entries.append(_build_entry(name_raw, tags, raw_lines))
                i = j
                continue

        i += 1

    return entries


def _build_entry(name_raw: str, tags: list[str], raw_lines: list[str]) -> dict:
    # Nome legível: \n → " / "
    name_clean = name_raw.replace("\\n", " / ")

    b64_lines = len(raw_lines)
    frames    = b64_lines // 4 if b64_lines >= 4 else 0

    if b64_lines == 0:
        status = "Rascunho"
    elif b64_lines == 4:
        status = "Completo"
    else:
        status = f"Multi-frame ({frames})"

    tag_set = set(tags)
    pos_tags  = sorted(tag_set & POSITION_TAGS)
    tech_tags = sorted(tag_set & TECHNIQUE_TAGS)
    top_pos   = sorted(tag_set & TOP_POSTURE_TAGS)
    bot_pos   = sorted(tag_set & BOTTOM_POSTURE_TAGS)

    main_cat, cat_detail = _categorize(tag_set, pos_tags, tech_tags, frames)

    # Decodificação 3D (todos os frames)
    decoded_frames = decode_all_frames(raw_lines) if raw_lines else None

    # Métricas resumidas do frame 0 (posição inicial)
    p0_cx = p0_cy = p0_cz = None
    p1_cx = p1_cy = p1_cz = None
    p0_height = p1_height = dist = None

    if decoded_frames:
        f0 = decoded_frames[0]
        cx0 = centroid(f0["player0"])
        cx1 = centroid(f0["player1"])
        p0_cx, p0_cy, p0_cz = cx0
        p1_cx, p1_cy, p1_cz = cx1
        p0_height = player_height(f0["player0"])
        p1_height = player_height(f0["player1"])
        dist = players_distance(f0["player0"], f0["player1"])

    return {
        "name_raw":        name_raw,
        "name_clean":      name_clean,
        "tags":            ", ".join(tags),
        "tag_count":       len(tags),
        "b64_lines":       b64_lines,
        "frames":          frames,
        "status":          status,
        "pos_tags":        ", ".join(pos_tags),
        "tech_tags":       ", ".join(tech_tags),
        "top_posture":     ", ".join(top_pos),
        "bottom_posture":  ", ".join(bot_pos),
        "main_category":   main_cat,
        "category_detail": cat_detail,
        # dados 3D brutos
        "decoded_frames":  decoded_frames,
        # métricas resumidas (frame 0)
        "p0_cx": p0_cx, "p0_cy": p0_cy, "p0_cz": p0_cz,
        "p1_cx": p1_cx, "p1_cy": p1_cy, "p1_cz": p1_cz,
        "p0_height": p0_height, "p1_height": p1_height,
        "dist_players": dist,
    }


def _categorize(tag_set, pos_tags, tech_tags, frames):
    """Infere categoria principal e detalhe a partir das tags."""
    # Submissões
    sub_tags = {"armbar", "kimura", "omoplata", "triangle", "rear_naked_choke",
                "arm_choke", "arm_triangle", "guillotine", "darce", "toehold",
                "heel_hook", "knee_bar", "neck_crank", "shoulder_lock", "monoplata"}
    if tag_set & sub_tags:
        return "Finalização", ", ".join(sorted(tag_set & sub_tags))

    # Quedas/takedowns
    td_tags = {"takedown", "double_leg_takedown", "single_leg_takedown",
               "throw", "sacrifice_throw", "te_waza", "koshi_waza", "ashi_waza", "sutemi_waza"}
    if tag_set & td_tags:
        return "Queda", ", ".join(sorted(tag_set & td_tags))

    # Passagens
    pass_tags = {t for t in tag_set if t.startswith("pass")}
    if pass_tags:
        return "Passagem de Guarda", ", ".join(sorted(pass_tags))

    # Raspagens
    if "sweep" in tag_set or "hip_bump" in tag_set:
        return "Raspagem", ", ".join(sorted(tag_set & {"sweep", "hip_bump", "butterfly_sweep"}))

    # Escapes / saídas
    if tag_set & {"stand_up", "bridge", "escape", "guard_pull", "guard_jump", "roll"}:
        return "Escape / Saída", ", ".join(sorted(tag_set & {"stand_up", "bridge", "escape", "guard_pull", "guard_jump", "roll"}))

    # Posição em pé
    if "standing" in tag_set and not frames:
        return "Em Pé", ", ".join(pos_tags)

    # Posição de guarda
    guard_tags = {t for t in tag_set if "guard" in t}
    if guard_tags:
        return "Guarda", ", ".join(sorted(guard_tags))

    # Posição de dominância
    dom_tags = {"mount", "side_control", "judo_side", "north_south", "back",
                "turtle", "knee_on_belly", "crucifix", "truck"}
    if tag_set & dom_tags:
        return "Posição Dominante", ", ".join(sorted(tag_set & dom_tags))

    # Posição específica
    if pos_tags:
        return "Posição", ", ".join(pos_tags)

    # Técnica genérica
    if tech_tags:
        return "Técnica", ", ".join(tech_tags)

    # Qualquer tag restante
    if tag_set:
        return "Outro", ", ".join(sorted(tag_set)[:3])

    return "Sem categoria", ""


# ── Estilos Excel ─────────────────────────────────────────────────────────────

def hex_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(name="Calibri", size=10)
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN    = Alignment(vertical="top")

THIN = Side(style="thin", color="D0D0D0")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CATEGORY_COLORS = {
    "Finalização":       "C0392B",
    "Queda":             "8E44AD",
    "Passagem de Guarda":"2980B9",
    "Raspagem":          "27AE60",
    "Escape / Saída":    "F39C12",
    "Em Pé":             "16A085",
    "Guarda":            "2471A3",
    "Posição Dominante": "922B21",
    "Posição":           "1E8449",
    "Técnica":           "6E2FBB",
    "Outro":             "717D7E",
    "Sem categoria":     "AAB7B8",
}

HEADER_FILLS = {cat: hex_fill(color) for cat, color in CATEGORY_COLORS.items()}
DEFAULT_HEADER_FILL = hex_fill("2C3E50")


def write_header(ws, columns: list[tuple[str, int]], fill: PatternFill = None):
    """Escreve linha de cabeçalho com estilo."""
    fill = fill or DEFAULT_HEADER_FILL
    for col_idx, (title, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font    = HEADER_FONT
        cell.fill   = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border  = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def write_row(ws, row_idx: int, values: list, category: str = ""):
    cat_color = CATEGORY_COLORS.get(category, "FFFFFF")
    row_fill_hex = _lighten(cat_color) if category else "FFFFFF"
    row_fill = hex_fill(row_fill_hex)
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font      = BODY_FONT
        cell.fill      = row_fill
        cell.alignment = TOP_ALIGN
        cell.border    = THIN_BORDER


def _lighten(hex_color: str, factor: float = 0.88) -> str:
    """Clareia uma cor hex em direção ao branco."""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"{r:02X}{g:02X}{b:02X}"


# ── Colunas ───────────────────────────────────────────────────────────────────

COLS_MAIN = [
    ("Nome",              30),
    ("Categoria",         20),
    ("Detalhe",           28),
    ("Tags",              50),
    ("Posição no Jogo",   24),
    ("Técnica",           22),
    ("Postura Top",       20),
    ("Postura Bottom",    20),
    ("Status",            14),
    ("Frames",             9),
    ("Qtd Tags",           9),
    # métricas 3D resumidas
    ("Lut.A Cx (m)",      11),
    ("Lut.A Cy (m)",      11),
    ("Lut.A Cz (m)",      11),
    ("Lut.B Cx (m)",      11),
    ("Lut.B Cy (m)",      11),
    ("Lut.B Cz (m)",      11),
    ("Alt.A (m)",         10),
    ("Alt.B (m)",         10),
    ("Dist. A↔B (m)",     13),
]

COLS_TAGS = [
    ("Tag",              28),
    ("Qtd Entradas",     14),
    ("Categorias Assoc.", 50),
]

# Colunas da aba de coordenadas 3D normalizadas
COLS_3D = (
    [("Nome", 32), ("Categoria", 20), ("Frame", 8), ("Lutador", 10)]
    + [(f"{j}_X", 9) for j in JOINTS_PT]
    + [(f"{j}_Y", 9) for j in JOINTS_PT]
    + [(f"{j}_Z", 9) for j in JOINTS_PT]
)

# Colunas da aba de articulações por posição (wide: uma linha por posição × lutador)
COLS_WIDE = (
    [("Nome", 32), ("Categoria", 20), ("Status", 12), ("Frame", 8), ("Lutador", 10)]
    + sum([[
        (f"{j}_X", 8),
        (f"{j}_Y", 8),
        (f"{j}_Z", 8),
    ] for j in JOINTS_PT], [])
)


def entry_row_main(e: dict) -> list:
    return [
        e["name_clean"],
        e["main_category"],
        e["category_detail"],
        e["tags"],
        e["pos_tags"],
        e["tech_tags"],
        e["top_posture"],
        e["bottom_posture"],
        e["status"],
        e["frames"] or "",
        e["tag_count"],
        e["p0_cx"], e["p0_cy"], e["p0_cz"],
        e["p1_cx"], e["p1_cy"], e["p1_cz"],
        e["p0_height"], e["p1_height"],
        e["dist_players"],
    ]


# ── Construção das abas ───────────────────────────────────────────────────────

def build_sheet_all(wb: Workbook, entries: list[dict]):
    ws = wb.active
    ws.title = "Todas"
    write_header(ws, COLS_MAIN)
    for i, e in enumerate(entries, start=2):
        write_row(ws, i, entry_row_main(e), e["main_category"])


def build_sheet_category(wb: Workbook, name: str, entries: list[dict], category: str):
    ws = wb.create_sheet(title=name)
    subset = [e for e in entries if e["main_category"] == category]
    subset.sort(key=lambda e: (e["category_detail"], e["name_clean"]))
    fill = HEADER_FILLS.get(category, DEFAULT_HEADER_FILL)
    write_header(ws, COLS_MAIN, fill=fill)
    for i, e in enumerate(subset, start=2):
        write_row(ws, i, entry_row_main(e), category)
    return len(subset)


def build_sheet_positions(wb: Workbook, entries: list[dict]):
    """Aba com TODAS as entradas frame=1 (posições fixas)."""
    ws = wb.create_sheet(title="Posições (frame único)")
    subset = [e for e in entries if e["frames"] == 1]
    subset.sort(key=lambda e: (e["main_category"], e["name_clean"]))
    write_header(ws, COLS_MAIN)
    for i, e in enumerate(subset, start=2):
        write_row(ws, i, entry_row_main(e), e["main_category"])


def build_sheet_transitions(wb: Workbook, entries: list[dict]):
    """Aba com entradas multi-frame (animações / transições)."""
    ws = wb.create_sheet(title="Técnicas Multi-frame")
    subset = [e for e in entries if e["frames"] > 1]
    subset.sort(key=lambda e: (e["frames"], e["name_clean"]))
    write_header(ws, COLS_MAIN)
    for i, e in enumerate(subset, start=2):
        write_row(ws, i, entry_row_main(e), e["main_category"])


def build_sheet_drafts(wb: Workbook, entries: list[dict]):
    """Aba com entradas sem dados (rascunhos)."""
    ws = wb.create_sheet(title="Rascunhos")
    subset = [e for e in entries if e["status"] == "Rascunho"]
    subset.sort(key=lambda e: e["name_clean"])
    write_header(ws, COLS_MAIN)
    for i, e in enumerate(subset, start=2):
        write_row(ws, i, entry_row_main(e), e["main_category"])


def build_sheet_3d_normalized(wb: Workbook, entries: list[dict]):
    """
    Aba normalizada: uma linha por (entrada × frame × lutador).
    Colunas: Nome, Categoria, Frame, Lutador, depois X/Y/Z de cada articulação
    como valores numéricos separados (uma coluna por coord).
    Só inclui entradas com dados 3D.
    """
    ws = wb.create_sheet(title="Coordenadas 3D")
    n_joints = len(JOINTS_PT)

    # Cabeçalho: Nome, Categoria, Frame, Lutador | então X de cada joint | Y | Z
    header = ["Nome", "Categoria", "Frame", "Lutador"]
    for axis in ("X", "Y", "Z"):
        for j in JOINTS_PT:
            header.append(f"{j}_{axis}")

    fill = DEFAULT_HEADER_FILL
    for col_idx, title in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font  = HEADER_FONT
        cell.fill  = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER

    # Larguras
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 10
    for col_idx in range(5, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 8

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "E2"

    row_idx = 2
    for e in entries:
        if not e["decoded_frames"]:
            continue
        for frame_num, frame in enumerate(e["decoded_frames"]):
            for player_key, player_label in (("player0", "Lutador A"), ("player1", "Lutador B")):
                coords = frame[player_key]  # list of (x,y,z) per joint
                row_data = [
                    e["name_clean"],
                    e["main_category"],
                    frame_num + 1,
                    player_label,
                ]
                # X de todos os joints
                for x, y, z in coords:
                    row_data.append(x)
                # Y de todos os joints
                for x, y, z in coords:
                    row_data.append(y)
                # Z de todos os joints
                for x, y, z in coords:
                    row_data.append(z)

                cat_color = CATEGORY_COLORS.get(e["main_category"], "FFFFFF")
                row_fill  = hex_fill(_lighten(cat_color))
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font   = BODY_FONT
                    cell.fill   = row_fill
                    cell.border = THIN_BORDER
                    cell.alignment = TOP_ALIGN
                row_idx += 1

    return row_idx - 2  # linhas gravadas


def build_sheet_3d_wide(wb: Workbook, entries: list[dict]):
    """
    Aba wide: uma linha por (entrada × frame × lutador) com cada joint em 3 colunas X/Y/Z.
    Layout mais amigável para análise direta no Excel.
    Inclui apenas o frame 0 (posição estática) para simplificar.
    """
    ws = wb.create_sheet(title="3D por Articulação")

    # Cabeçalho
    header_meta = ["Nome", "Categoria", "Status", "Lutador"]
    header_joints = []
    for j in JOINTS_PT:
        header_joints += [f"{j} X", f"{j} Y", f"{j} Z"]
    header = header_meta + header_joints

    fill = DEFAULT_HEADER_FILL
    for col_idx, title in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font  = HEADER_FONT
        cell.fill  = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    for col_idx in range(5, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 7

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "E2"

    row_idx = 2
    for e in entries:
        if not e["decoded_frames"]:
            continue
        f0 = e["decoded_frames"][0]
        for player_key, player_label in (("player0", "Lutador A"), ("player1", "Lutador B")):
            coords = f0[player_key]
            row_data = [
                e["name_clean"],
                e["main_category"],
                e["status"],
                player_label,
            ]
            for x, y, z in coords:
                row_data += [x, y, z]

            cat_color = CATEGORY_COLORS.get(e["main_category"], "FFFFFF")
            row_fill  = hex_fill(_lighten(cat_color))
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font   = BODY_FONT
                cell.fill   = row_fill
                cell.border = THIN_BORDER
                cell.alignment = TOP_ALIGN
            row_idx += 1

    return row_idx - 2


def build_sheet_tags(wb: Workbook, entries: list[dict]):
    ws = wb.create_sheet(title="Tags")
    # Conta uso de cada tag
    tag_counter: Counter = Counter()
    tag_categories: dict[str, set] = {}
    for e in entries:
        for t in e["tags"].split(", ") if e["tags"] else []:
            tag_counter[t] += 1
            tag_categories.setdefault(t, set()).add(e["main_category"])

    write_header(ws, COLS_TAGS)
    for i, (tag, count) in enumerate(tag_counter.most_common(), start=2):
        cats = ", ".join(sorted(tag_categories.get(tag, set())))
        ws.cell(row=i, column=1, value=tag).font = BODY_FONT
        ws.cell(row=i, column=2, value=count).font = BODY_FONT
        ws.cell(row=i, column=3, value=cats).font = BODY_FONT
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = THIN_BORDER
            ws.cell(row=i, column=col).alignment = TOP_ALIGN


def build_sheet_summary(wb: Workbook, entries: list[dict]):
    ws = wb.create_sheet(title="Resumo", index=0)  # primeira aba
    ws.sheet_view.showGridLines = False

    title_font  = Font(name="Calibri", bold=True, size=16, color="1B2631")
    label_font  = Font(name="Calibri", bold=True, size=11)
    value_font  = Font(name="Calibri", size=11)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16

    ws["A1"] = "GrappleMap — Exportação"
    ws["A1"].font = title_font
    ws.row_dimensions[1].height = 30

    ws["A2"] = "Fonte: github.com/Eelis/GrappleMap"
    ws["A2"].font = Font(name="Calibri", size=10, color="717D7E", italic=True)

    ws.append([])

    # totais
    total           = len(entries)
    complete        = sum(1 for e in entries if e["status"] == "Completo")
    multi           = sum(1 for e in entries if e["frames"] > 1)
    drafts          = sum(1 for e in entries if e["status"] == "Rascunho")
    submissions     = sum(1 for e in entries if e["main_category"] == "Finalização")
    takedowns       = sum(1 for e in entries if e["main_category"] == "Queda")
    passes          = sum(1 for e in entries if e["main_category"] == "Passagem de Guarda")
    sweeps          = sum(1 for e in entries if e["main_category"] == "Raspagem")
    guards          = sum(1 for e in entries if e["main_category"] == "Guarda")
    dominant        = sum(1 for e in entries if e["main_category"] == "Posição Dominante")

    rows = [
        ("Total de entradas",       total),
        ("Posições completas",       complete),
        ("Técnicas multi-frame",     multi),
        ("Rascunhos / incompletas",  drafts),
        ("", ""),
        ("Finalizações",             submissions),
        ("Quedas / Takedowns",       takedowns),
        ("Passagens de Guarda",      passes),
        ("Raspagens",                sweeps),
        ("Guardas",                  guards),
        ("Posições Dominantes",      dominant),
    ]

    for label, value in rows:
        row = ws.max_row + 1
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.font = label_font
        vc.font = value_font
        vc.alignment = Alignment(horizontal="right")

    # gráfico de categorias
    cat_counts = Counter(e["main_category"] for e in entries)
    ws.append([])
    ws.append(["Categoria", "Qtd"])
    ws["A" + str(ws.max_row)].font = Font(bold=True)
    ws["B" + str(ws.max_row)].font = Font(bold=True)
    for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=cat).font = value_font
        wc = ws.cell(row=row, column=2, value=count)
        wc.font = value_font
        wc.alignment = Alignment(horizontal="right")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Baixando GrappleMap.txt…")
    resp = requests.get(URL, timeout=30)
    resp.raise_for_status()
    text = resp.text
    print(f"  {len(text):,} bytes baixados")

    print("Parseando entradas…")
    entries = parse_grapplemap(text)
    print(f"  {len(entries)} entradas encontradas")

    cat_counts = Counter(e["main_category"] for e in entries)
    for cat, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
        print(f"    {cat:<28} {n:>4}")

    print(f"\nGerando {OUTPUT}…")
    wb = Workbook()
    # Remove aba padrão temporariamente (será recriada como "Todas")
    default = wb.active
    wb.remove(default)

    build_sheet_summary(wb, entries)

    # Aba geral
    ws_all = wb.create_sheet(title="Todas")
    write_header(ws_all, COLS_MAIN)
    for i, e in enumerate(entries, start=2):
        write_row(ws_all, i, entry_row_main(e), e["main_category"])

    # Abas por categoria
    category_sheets = [
        ("Finalizações",        "Finalização"),
        ("Quedas",              "Queda"),
        ("Passagens",           "Passagem de Guarda"),
        ("Raspagens",           "Raspagem"),
        ("Pos. Dominantes",     "Posição Dominante"),
        ("Guardas",             "Guarda"),
        ("Em Pé",               "Em Pé"),
        ("Escapes",             "Escape / Saída"),
        ("Posições",            "Posição"),
        ("Técnicas",            "Técnica"),
        ("Outros",              "Outro"),
    ]
    for sheet_name, category in category_sheets:
        n = build_sheet_category(wb, sheet_name, entries, category)
        if n == 0:
            del wb[sheet_name]

    # Abas especiais
    build_sheet_positions(wb, entries)
    build_sheet_transitions(wb, entries)
    build_sheet_drafts(wb, entries)
    build_sheet_tags(wb, entries)

    # Abas 3D
    print("Gerando abas 3D…")
    n_wide = build_sheet_3d_wide(wb, entries)
    print(f"  3D por Articulação: {n_wide} linhas (frame 0 × 2 lutadores)")
    n_norm = build_sheet_3d_normalized(wb, entries)
    print(f"  Coordenadas 3D (todos frames): {n_norm} linhas")

    wb.save(OUTPUT)
    print(f"  Salvo: {OUTPUT}")
    print(f"  Abas: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    main()
